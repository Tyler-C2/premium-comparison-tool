from openpyxl import *
from openpyxl.styles import Border, Side, PatternFill, Fill, Font, GradientFill,Alignment
from openpyxl.utils import cell

class Parser():
    def format_date(start_date, end_date):
        start_formatted = f"{start_date.month}/{start_date.day}/{start_date.year}"
        end_formatted = f"{end_date.month}/{end_date.day}/{end_date.year}"
        
        return (start_formatted,end_formatted)
    
    # used to check if input value can be converted to a float
    def floatCheck(val1,val2):
        vals = [val1,val2]
        for i in range(len(vals)):
            try:
                float(vals[i])
                vals[i] = True
            except ValueError:
                vals[i] = False
            
        return vals

    # line item processing
    def parse_line_items(c1_items, c2_items):
        car1_items = []
        car2_items = [] 

        for i in range(1,9):
            current_c1 = c1_items[f'c1_item{i}'].get()
            current_c2 = c2_items[f'c2_item{i}'].get()
            c1_items[f'c1_item{i}'].set("")
            c2_items[f'c2_item{i}'].set("")  
            
            checks = Parser.floatCheck(current_c1, current_c2)
            
            if checks[0] == True:
                car1_items.append(round(float(current_c1),2))
            else:
                car1_items.append("n/a")
            
            if checks[1] == True:
                car2_items.append(round(float(current_c2),2))
            else:
                car2_items.append("n/a")

        return (car1_items, car2_items)

# class to create instances of periods
class PeriodData():
    def __init__(self, date, cars, car1_premium, car2_premium):
        self.date = date
        self.cars = cars
        self.car1_premium = car1_premium
        self.car2_premium = car2_premium

class WorkbookCreator():
    def __init__(self):
        self.path = None
        self.file_name = None
        self.styles()
        self.items = self.line_items()
        self.period_data_lst = []

    def difference_sheet(self):
        if "Difference" in self.wb.sheetnames:
            return self.wb["Difference"]
        else:
            return self.wb.create_sheet(title="Difference")

    def styles(self):
        self.white_fill = PatternFill(start_color="FFFFFF",end_color="FFFFFF",fill_type=None)
        self.alt_fill = PatternFill(start_color="EDECE9",end_color="EDECE9",fill_type="solid") 
        self.pos_fill = PatternFill(start_color="7FBF7F",end_color="7FBF7F",fill_type="solid") 
        self.neg_fill = PatternFill(start_color="BF7F7F",end_color="BF7F7F",fill_type="solid") 
        self.neutral_fill = PatternFill(start_color="BF7F7F",end_color="BF7F7F",fill_type="solid")
        self.center = Alignment(horizontal='center',vertical='center')
        self.new_border = Border(left=Side(border_style='thin', color='969696'),
                            right=Side(border_style='thin', color='969696'),
                            top=Side(border_style='thin', color='969696'),
                            bottom=Side(border_style='thin', color='969696'),
                            )

    def line_items(self):
        line_items = [
            "Bodily Injury","Property Damage","Medical Payments",
            "Unisured Motorist Option","Unisured Motorist Bodily Injury",
            "Unisured Motorist Property Damage","Other Than Collision","Collision"
        ]
        
        return line_items

    def add_path(self, file_path):
        supported = ["xlsx","xlsm","xltx","xltm"]

        if file_path[-4:] in supported:
            self.path = file_path
            self.get_file_name()
        else:
            self.path == self.path

    def get_file_name(self):
        path_end_idx=self.path.rfind('/')
        self.file_name = self.path[path_end_idx+1:]

    # creates class that holds period content
    def create_period(self, date, cars, car1_premium, car2_premium):
        new_period = PeriodData(date, cars, car1_premium, car2_premium)
        self.period_data_lst.append(new_period)

    # adds all current period content to workbook 
    def finished_input(self,path):
        self.wb = load_workbook(path)
        self.item_ws = self.wb.active
        self.dif_ws = self.difference_sheet()

        for p in self.period_data_lst:
            self.row_names(self.item_ws)
            self.populate_data(p.date, p.cars, p.car1_premium, p.car2_premium)
            self.row_names(self.dif_ws)
            self.calc_dif(p.cars)
            self.save_workbook()

        self.period_data_lst.clear()

    def difference_sheet(self):
        if "Difference" in self.wb.sheetnames:
            return self.wb["Difference"]
        else:
            return self.wb.create_sheet(title="Difference")

    def row_names(self,cur_ws):
        cur_ws.column_dimensions["A"].width = 20

        i = 0
        for row in range(3,11):
            cur_ws.cell(column=1, row=row, value=self.items[i])
            i += 1
            cur_ws['A'+str(row)].alignment = Alignment(wrap_text=True)

    def populate_data(self, date, cars, car1_premium, car2_premium):
        current_cell = self.item_ws.cell(column=2, row=1)
        
        while current_cell.value: # finds empty col
            current_cell = self.item_ws.cell(column = current_cell.column+2, row = 1)
            
        col1 = current_cell.column
        col2 = current_cell.column + 1

        self.date_entry(col1,date)
        self.col_entry(col1, cars[0], car1_premium)
        self.col_entry(col2, cars[1], car2_premium)

        self.format_item_entry(col1)

    def format_item_entry(self, col):
        date = self.item_ws.cell(column=col,row=1)

        if col <= 3:
            group_fill = self.alt_fill
        elif self.item_ws.cell(column=col-2,row=1).fill.fill_type == None:
            group_fill = self.alt_fill
        else:
            group_fill = self.white_fill

        date.alignment = self.center
        date.fill = group_fill
        date.border= self.new_border

        col_letter = cell.get_column_letter(col)
        self.item_ws.column_dimensions[col_letter].width = 12

        for i in range(2):
            current_cell = self.item_ws.cell(column=col+i,row=2)
            while current_cell.value:
                current_cell.alignment = self.center
                current_cell.fill = group_fill
                current_cell.border = self.new_border
                current_cell = self.item_ws.cell(column=col+i, row = current_cell.row+1)

    def date_entry(self, col, date):
        self.item_ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        self.item_ws.cell(column = col, row = 1, value = f"{date[0]} - {date[1]}")

    def col_entry(self, col, car, premium):
        row = 3

        self.item_ws.cell(column = col, row = 2, value = car)

        for val in premium:
            self.item_ws.cell(column = col, row = row, value = val)
            row +=1

    def calc_dif(self, cars):
        active_cell = self.item_ws.cell(column=2, row=1)
        
        while active_cell.value: 
            #date merge and populate
            self.dif_ws.merge_cells(start_row=1, start_column=active_cell.column, end_row=1, end_column=active_cell.column+1)
            self.dif_ws.cell(column = active_cell.column, row = active_cell.row, value = active_cell.value)
            
            date_letter = cell.get_column_letter(active_cell.column)
            self.dif_ws.column_dimensions[date_letter].width = 12
            
            self.dif_ws.cell(column = active_cell.column,row=1).alignment = self.center

            self.dif_ws.cell(column = active_cell.column, row = 2, value = cars[0]).alignment = self.center
            self.col_dif(active_cell.column)

            self.dif_ws.cell(column = active_cell.column+1, row = 2, value = cars[1]).alignment = self.center
            self.col_dif(active_cell.column+1)

            active_cell = self.item_ws.cell(column = active_cell.column+2, row = 1)

    def col_dif(self, col):
        active_cell = self.item_ws.cell(column = col, row = 3)

        if col > 3:
            for i in range(8):
                difference = 0
                prev_cell = self.item_ws.cell(column = active_cell.column-2, row = active_cell.row)
                
                active_check = self.cell_check(active_cell.value)
                prev_check = self.cell_check(prev_cell.value)
                
                if active_check == True and prev_check == True:
                    difference = float(active_cell.value)-float(prev_cell.value)
                    dif_cell = self.dif_ws.cell(column = active_cell.column, row = active_cell.row, value = difference)
                else:
                    dif_cell = self.dif_ws.cell(column = active_cell.column, row = active_cell.row, value = "n/a")

                self.dif_sheet_style(dif_cell,difference)

                active_cell = self.item_ws.cell(column = col, row = active_cell.row+1)

    def dif_sheet_style(self,cell,dif):
        cell.alignment = self.center
        cell.border= self.new_border

        if dif < 0:
            cell.fill = self.pos_fill
        elif dif > 0:
            cell.fill = self.neg_fill
        else:
            cell.fill = self.alt_fill

    def cell_check(self, cell_value):
        if cell_value != "n/a":
            return True
        else:
            return False


    def save_workbook(self):
        self.wb.save(self.path)
